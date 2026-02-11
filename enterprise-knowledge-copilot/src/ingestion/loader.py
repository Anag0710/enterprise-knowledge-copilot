from pathlib import Path
from typing import List, Dict
import pdfplumber

# Multi-format support
try:
    from docx import Document as DocxDocument
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from bs4 import BeautifulSoup
    import markdown
    HTML_AVAILABLE = True
except ImportError:
    HTML_AVAILABLE = False

# Import chunker correctly
from src.ingestion.chunker import chunk_text
from src.embeddings.vector_store import VectorStore


def clean_text(text: str) -> str:
    """Basic text cleaning for enterprise documents."""
    text = text.replace("\n", " ")
    text = " ".join(text.split())
    return text


def load_pdf(file_path: Path) -> List[Dict]:
    """
    Load a PDF file and extract text page by page.
    """
    documents = []

    with pdfplumber.open(file_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text()
            if not text:
                continue

            documents.append({
                "text": clean_text(text),
                "metadata": {
                    "source": file_path.name,
                    "page": page_num
                }
            })

    return documents


def load_docx(file_path: Path) -> List[Dict]:
    """Load a Word document (.docx)."""
    if not DOCX_AVAILABLE:
        raise ImportError("python-docx not installed. Run: pip install python-docx")
    
    doc = DocxDocument(file_path)
    documents = []
    
    # Group paragraphs into pages (roughly 500 words per page)
    current_page = []
    current_word_count = 0
    page_num = 1
    
    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        
        word_count = len(text.split())
        current_page.append(text)
        current_word_count += word_count
        
        # Split into "pages" every 500 words
        if current_word_count >= 500:
            documents.append({
                "text": clean_text(" ".join(current_page)),
                "metadata": {
                    "source": file_path.name,
                    "page": page_num
                }
            })
            current_page = []
            current_word_count = 0
            page_num += 1
    
    # Add remaining text
    if current_page:
        documents.append({
            "text": clean_text(" ".join(current_page)),
            "metadata": {
                "source": file_path.name,
                "page": page_num
            }
        })
    
    return documents


def load_excel(file_path: Path) -> List[Dict]:
    """Load an Excel spreadsheet (.xlsx)."""
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")
    
    wb = load_workbook(file_path, read_only=True, data_only=True)
    documents = []
    
    for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
        sheet = wb[sheet_name]
        rows_text = []
        
        for row in sheet.iter_rows(values_only=True):
            # Filter out None values and convert to strings
            row_values = [str(cell) for cell in row if cell is not None]
            if row_values:
                rows_text.append(" | ".join(row_values))
        
        if rows_text:
            documents.append({
                "text": clean_text(" ".join(rows_text)),
                "metadata": {
                    "source": file_path.name,
                    "page": sheet_idx,
                    "sheet": sheet_name
                }
            })
    
    return documents


def load_text(file_path: Path) -> List[Dict]:
    """Load a plain text file (.txt)."""
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read()
    
    # Split into ~1000 word chunks as "pages"
    words = text.split()
    chunks = []
    page_num = 1
    
    for i in range(0, len(words), 1000):
        chunk_words = words[i:i+1000]
        if chunk_words:
            chunks.append({
                "text": clean_text(" ".join(chunk_words)),
                "metadata": {
                    "source": file_path.name,
                    "page": page_num
                }
            })
            page_num += 1
    
    return chunks if chunks else [{"text": clean_text(text), "metadata": {"source": file_path.name, "page": 1}}]


def load_markdown(file_path: Path) -> List[Dict]:
    """Load a Markdown file (.md) and convert to plain text."""
    if not HTML_AVAILABLE:
        # Fallback to plain text loading
        return load_text(file_path)
    
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        md_text = f.read()
    
    # Convert markdown to HTML, then extract text
    html = markdown.markdown(md_text)
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text(separator=" ", strip=True)
    
    # Split into pages similar to text files
    words = text.split()
    chunks = []
    page_num = 1
    
    for i in range(0, len(words), 1000):
        chunk_words = words[i:i+1000]
        if chunk_words:
            chunks.append({
                "text": clean_text(" ".join(chunk_words)),
                "metadata": {
                    "source": file_path.name,
                    "page": page_num
                }
            })
            page_num += 1
    
    return chunks if chunks else [{"text": clean_text(text), "metadata": {"source": file_path.name, "page": 1}}]


def load_html(file_path: Path) -> List[Dict]:
    """Load an HTML file and extract text."""
    if not HTML_AVAILABLE:
        raise ImportError("beautifulsoup4 not installed. Run: pip install beautifulsoup4")
    
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        html_content = f.read()
    
    soup = BeautifulSoup(html_content, "html.parser")
    
    # Remove script and style elements
    for script in soup(["script", "style"]):
        script.decompose()
    
    text = soup.get_text(separator=" ", strip=True)
    
    # Split into pages
    words = text.split()
    chunks = []
    page_num = 1
    
    for i in range(0, len(words), 1000):
        chunk_words = words[i:i+1000]
        if chunk_words:
            chunks.append({
                "text": clean_text(" ".join(chunk_words)),
                "metadata": {
                    "source": file_path.name,
                    "page": page_num
                }
            })
            page_num += 1
    
    return chunks if chunks else [{"text": clean_text(text), "metadata": {"source": file_path.name, "page": 1}}]


def load_document(file_path: Path) -> List[Dict]:
    """
    Auto-detect file type and load appropriately.
    Supports: PDF, DOCX, XLSX, TXT, MD, HTML
    """
    suffix = file_path.suffix.lower()
    
    loaders = {
        ".pdf": load_pdf,
        ".docx": load_docx,
        ".xlsx": load_excel,
        ".txt": load_text,
        ".md": load_markdown,
        ".markdown": load_markdown,
        ".html": load_html,
        ".htm": load_html,
    }
    
    loader = loaders.get(suffix)
    if not loader:
        print(f"Warning: Unsupported file type '{suffix}' for {file_path.name}")
        return []
    
    try:
        return loader(file_path)
    except Exception as e:
        print(f"Error loading {file_path.name}: {e}")
        return []


def load_documents_from_directory(directory: Path, pattern: str = "*.*") -> List[Dict]:
    """
    Load all supported documents inside the directory.
    Supports: PDF, DOCX, XLSX, TXT, MD, HTML
    """
    all_documents: List[Dict] = []
    supported_extensions = {".pdf", ".docx", ".xlsx", ".txt", ".md", ".markdown", ".html", ".htm"}
    
    if pattern == "*.pdf":
        # Backward compatibility
        pattern = "*.*"
    
    for file_path in sorted(directory.glob(pattern)):
        if file_path.suffix.lower() in supported_extensions:
            docs = load_document(file_path)
            all_documents.extend(docs)
    
    return all_documents

if __name__ == "__main__":
    docs_path = Path("data/raw_docs")
    pages = load_documents_from_directory(docs_path)
    chunks = chunk_text(pages)

    store = VectorStore()
    store.build_index(chunks)

    query = "What is the leave policy?"
    results = store.search(query)

    print("Top retrieved chunk:")
    print(results[0])
